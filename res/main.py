import pandas as pd
import re
from openpyxl.utils import get_column_letter
from pandas import DataFrame
from openpyxl import load_workbook
from utils import load_excel

# ---------- Utility: Normalize Mobile Numbers ----------
def normalize_mobile(num):
    """Clean and standardize mobile numbers (keep last 10 digits)"""
    if pd.isna(num):
        return None
    num = str(num)
    num = re.sub(r'\D', '', num)  # remove all non-digits
    num = re.sub(r'^(91|0|91)?', '', num)  # remove leading +91, 91, or 0
    if len(num) > 10:
        num = num[-10:]  # keep last 10 digits
    return num if len(num) == 10 else None

# ---------- Utility: Detect Possible Phone Columns ----------
def get_phone_columns(df: DataFrame):
    """Find columns that likely contain phone numbers by dynamically detecting the header row"""
    header_keywords = ['s.no', 'name', 'phone', 'mobile', 'ph', 'contact', 'number', 'ph.no', 'ID', 'Name', 'Email', 'Phone', 'Status']
    phone_keywords = ['phone', 'Phone', 'mobile', 'ph', 'contact', 'number', 'PH.NO.', 'Ph.no']
    
    def is_header_row(row):
        row_values = row.astype(str).str.lower()
        return any(any(keyword.lower() == str(val).lower() for keyword in header_keywords) for val in row_values)
    
    header_row_idx = None
    for idx, row in df.iterrows():
        if is_header_row(row):
            header_row_idx = idx
            break
    
    if header_row_idx is not None:
        df.columns = df.iloc[header_row_idx].str.strip()
        df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    
    cols = [col for col in df.columns if any(k.lower() in str(col).lower() for k in phone_keywords)]
    return cols

# ---------- Load Files ----------
user_file = input('Give user file(.csv): ')
elec_file = input("Give electrician excel(.xls): ")
user_file = user_file.replace("\"", "")
elec_file = elec_file.replace("\"", "")

try:
    user_df = pd.read_csv(user_file, dtype=str)
except Exception:
    user_df = pd.read_excel(user_file, dtype=str)

# Load all sheets from electrician Excel file
xl = pd.ExcelFile(elec_file)
elec_dfs = {sheet: pd.read_excel(elec_file, sheet_name=sheet, dtype=str, header=1) for sheet in xl.sheet_names}
elec_wb = load_excel(elec_file, read_only=False)

# ---------- Extract and Normalize Numbers ----------
user_cols = get_phone_columns(user_df)
user_numbers = set()
for col in user_cols:
    user_numbers.update(user_df[col].map(normalize_mobile).dropna())

print(f"✅ Found {len(user_numbers)} unique mobile numbers in user file.")

# ---------- Build lookup for user.csv numbers with their cell address ----------
user_number_map = {}
for col_name in user_cols:
    col_index = user_df.columns.get_loc(col_name)
    for row_idx, val in enumerate(user_df[col_name]):
        num = normalize_mobile(val)
        if num:
            user_number_map[num] = f"{get_column_letter(col_index + 1)}{row_idx + 2}"

# ---------- Match Numbers Across All Sheets in Electrician File ----------
matches = []
print(f"{'Sr.':<5} | {'NAME':<25} | {'MOBILE':<12} | {'ELECTRICIAN':<30} | {'USER EXCEL':<8}")
print("-" * 90)

for sheet_name, elec_df in elec_dfs.items():
    elec_cols = get_phone_columns(elec_df)
    elec_ws = elec_wb[sheet_name]
    
    for row_idx, row in elec_df.iterrows():
        for col_name in elec_cols:
            num = normalize_mobile(row[col_name])
            if num and num in user_number_map:
                elec_cell = f"{sheet_name} ! {get_column_letter(elec_df.columns.get_loc(col_name)+1)}{row_idx+3}"
                user_cell = user_number_map[num]
                
                print('row number', row_idx)
                elec_ws[f'F{row_idx+3}'] = '    App install'
                matches.append({
                    "Sr No": row.get("Sr No", row_idx + 1),
                    "Name": row.get("Name", ""),
                    "Matched_Number": num,
                    "Electrician_Cell": elec_cell,
                    "User_Cell": user_cell
                })

                print(f"{row.get('Sr No', row_idx + 1):<5} | {row.get('NAME', ''):<25} | {num:<12} | "
                      f"{elec_cell:<30} | user → {user_cell}")
                print("-" * 90)
                
            

# ---------- Export Results ----------
if matches:
    result_df = pd.DataFrame(matches)
    result_df.to_excel("matched_numbers.xlsx", index=False)
    
    elec_wb.save(elec_file)
    print(f"\n✅ Matching complete! {len(matches)} matches found.")
    print("📄 Results saved to matched_numbers.xlsx")
else:
    print("\n❌ No matching numbers found.")
    
input("PRess enter to close: ")